from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from config.form_expectations import canonicalize_url
from config.url_loader import URL_FILE_BY_BRAND


FORM_MATRIX_COLS: dict[str, tuple[int, int]] = {
    "mts": (1, 2),
    "rostelecom": (3, 4),
    "beeline": (5, 6),
    "megafon": (7, 8),
    "domru": (9, 10),
    "t2": (11, 12),
    "ttk": (14, 15),
}

URL_PATTERN = re.compile(r"https?://[^\s\])]+", flags=re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate URL form-overrides from xlsx sheet 'Тесткейсы форм'.",
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        help="Path to workbook 'Автотесты Лендинги.xlsx'.",
    )
    parser.add_argument(
        "--project-root",
        default=".",
        help="Repository root with urls/ and config/ directories.",
    )
    return parser.parse_args()


def load_urls_by_brand(urls_dir: Path) -> dict[str, set[str]]:
    result: dict[str, set[str]] = {}
    for brand, file_name in URL_FILE_BY_BRAND.items():
        file_path = urls_dir / file_name
        values: set[str] = set()
        for raw in file_path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            values.add(canonicalize_url(line))
        result[brand] = values
    return result


def extract_url(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    match = URL_PATTERN.search(text)
    if not match:
        return None
    return canonicalize_url(match.group(0))


def parse_note(note_raw: object) -> tuple[list[str], list[str]]:
    note = str(note_raw or "").strip().lower()
    if not note:
        return [], []

    missing: set[str] = set()
    force_present: set[str] = set()

    # missing forms by explicit markers
    if re.search(r"нет\s+checkaddress-undecided\b", note):
        missing.add("undecided")
    if re.search(r"нет\s+checkaddress_address_button\b", note):
        missing.add("checkaddress")
    if re.search(r"нет\s+checkaddress\b(?![-_])", note):
        missing.add("checkaddress")
    if re.search(r"нет\s+формы?\s+profit\b", note) or re.search(r"нет\s+profit\b", note):
        missing.add("profit")
    if re.search(r"нет\s+business_no_address_button\b", note):
        missing.add("business")
    if re.search(r"нет\s+moving_address_button\b", note):
        missing.add("moving")
    if re.search(r"нет\s+express-connection_address_button\b", note):
        missing.add("express-connection")

    # positive markers
    if re.search(r"\bbusiness\b", note) and "business" not in missing:
        force_present.add("business")
    if re.search(r"\bmoving\b", note) and "moving" not in missing:
        force_present.add("moving")

    return sorted(missing), sorted(force_present)


def pick_form_matrix_sheet(workbook):
    for ws in workbook.worksheets:
        a1 = ws.cell(1, 1).value
        d1 = ws.cell(1, 4).value
        if str(a1 or "").strip() == "Форма" and str(d1 or "").strip() == "DataTestID":
            return ws
    raise RuntimeError("Form matrix sheet not found (expected A1='Форма', D1='DataTestID').")


def main() -> None:
    args = parse_args()
    project_root = Path(args.project_root).resolve()
    urls_dir = (project_root / "urls").resolve()
    output_dir = (project_root / "config" / "form_expectations").resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    urls_by_brand = load_urls_by_brand(urls_dir)
    wb = load_workbook(args.xlsx, data_only=True)
    ws = pick_form_matrix_sheet(wb)

    overrides_by_brand: dict[str, dict[str, dict[str, list[str]]]] = {
        brand: {} for brand in FORM_MATRIX_COLS
    }

    for row in range(45, ws.max_row + 1):
        for brand, (url_col, note_col) in FORM_MATRIX_COLS.items():
            url = extract_url(ws.cell(row, url_col).value)
            if not url:
                continue
            if url not in urls_by_brand.get(brand, set()):
                continue

            missing, force_present = parse_note(ws.cell(row, note_col).value)
            if not missing and not force_present:
                continue

            existing = overrides_by_brand[brand].setdefault(
                url,
                {"missing": [], "force_present": []},
            )
            existing["missing"] = sorted(set(existing["missing"]).union(missing))
            existing["force_present"] = sorted(
                set(existing["force_present"]).union(force_present)
            )

    for brand, data in overrides_by_brand.items():
        out_path = output_dir / f"{brand}.json"
        rendered = json.dumps(
            {url: rules for url, rules in sorted(data.items())},
            ensure_ascii=False,
            indent=2,
        )
        out_path.write_text(rendered + "\n", encoding="utf-8")
        print(f"[OK] {brand}: {len(data)} override URLs -> {out_path}")


if __name__ == "__main__":
    main()

